import shelve
# from pymodbus.payload import BinaryPayloadBuilder
from datetime import datetime
from openpyxl import Workbook, load_workbook
import sqlite3
import json

temp_date = 23
start_minute = 8
stop_minute = start_minute + 1

soligorsk_power = 'soligorsk_power'
soligorsk_water_cold = 'soligorsk_water_cold'
soligorsk_water_hot = 'soligorsk_water_hot'
soligorsk_heat = 'soligorsk_heat'

borisov_power_1 = 'borisov_power_1'
borisov_power_2 = 'borisov_power_2'
borisov_water = 'borisov_water'
borisov_heat = 'borisov_heat'

bobruisk_power = 'bobruisk_power'
bobruisk_water = 'bobruisk_water'
bobruisk_heat = 'bobruisk_heat'

minsk_co_gaz_1VR = 'minsk_co_gaz_1VR'
minsk_co_gaz_1VS = 'minsk_co_gaz_1VS'
minsk_co_gaz_1P = 'minsk_co_gaz_1P'
minsk_co_gaz_1T = 'minsk_co_gaz_1T'

dictionary_registers = {soligorsk_power: 0, soligorsk_water_cold: 1,
                        soligorsk_water_hot: 2, soligorsk_heat: 3,

                        borisov_power_1: 4, borisov_power_2: 5, borisov_water: 6, borisov_heat: 7,

                        bobruisk_power: 8, bobruisk_water: 9, bobruisk_heat: 10,

                        minsk_co_gaz_1VR: 11, minsk_co_gaz_1VS: 12, minsk_co_gaz_1P: 13, minsk_co_gaz_1T: 14}
dictionary_temp = {}
filename_shelve = 'db'


def gas(
        gas_year=datetime.today().year,
        gas_month=datetime.today().month,
        gas_day=datetime.today().day,
        gas_hour=datetime.today().hour,
        gas_minute=datetime.today().minute):
    try:
        wb = load_workbook('./gas/example_gas.xlsx')
        ws = wb.active
        print(ws)
        with shelve.open(filename_shelve) as states:
            # запись месяц назад на 1 число 9.00 ( 1, 9,)
            data_dictionary = states.get(datetime(gas_year, gas_month, gas_day, gas_hour,
                                                  gas_minute).strftime("%d%m%y %H:%M"), {0})
            print('ok data')
            # data_dictionary = states.get(data_save, {0})
            if data_dictionary == {0}:
                print('Для газа сохранения по этому адресу нет - 0')
                ws['C5'].value = 0
                ws['D5'].value = 0
                ws['E5'].value = 0
                ws['F5'].value = 0
                ws['H5'].value = 0
                ws['I5'].value = 0
            else:
                print('Для газа заполняем таблицу')
                ws['C5'].value = data_dictionary[minsk_co_gaz_1VS]
                ws['D5'].value = data_dictionary[minsk_co_gaz_1VR]
                ws['E5'].value = data_dictionary[minsk_co_gaz_1P]
                ws['F5'].value = data_dictionary[minsk_co_gaz_1T]
                ws['H5'].value = data_dictionary[minsk_co_gaz_1VS]
                ws['I5'].value = data_dictionary[minsk_co_gaz_1VR]
                # print(data_dictionary[minsk_co_gaz_1VS])

            data_dictionary.clear()
            states.close()

            wb.save(f'./gas/{datetime.today().strftime("%B%Y")}_gas.xlsx')  # сохраним в новой книге
            wb.close()
            states.close()

    except BaseException as e:
        print(f'ошибка   {e}')


def sqlite_copy_shelve(name_shelve, name_db, name_table):
    """
    :param name_shelve:
    :param name_db:
    :param name_table:
    :return:
    """
    base_conn = None
    try:
        base_conn = sqlite3.connect(name_db)
        cursor = base_conn.cursor()
        cursor.execute(f'CREATE TABLE IF NOT EXISTS {name_table}(data PRIMARY KEY, temp )')
        base_conn.commit()

        with shelve.open(name_shelve) as states:
            for key in states:
                cursor.execute(f'INSERT INTO {name_table} VALUES(?, ?)', (key, json.dumps(states[key])))
                base_conn.commit()
        states.close()
        print('База заполнена успешно.')

    except sqlite3.Error as error:
        print('Ошибка соединения с БД : ', error)
    finally:
        if base_conn:
            base_conn.close()


def sqlite_search_data(name_data, name_db, name_table):
    """
    :param name_data:
    :param name_db:
    :param name_table:
    :return: поиск данных по дате
    """

    base_conn = None
    try:
        base_conn = sqlite3.connect(name_db)
        cursor = base_conn.cursor()
        temp = cursor.execute(f'SELECT temp FROM {name_table} WHERE data == ?', (name_data,)).fetchone()
        print('Данные найдены успешно. ')
        # temp_list = json.loads(str(temp))
        # return temp_list
        return str(temp)

    except sqlite3.Error as error:
        print('Ошибка соединения с БД : ', error)
    finally:
        if base_conn:
            base_conn.close()


def print_shelve():
    with shelve.open(filename_shelve) as states:
        for key in states:
            print(key, " - ", states[key])
    states.close()


if __name__ == '__main__':
    a = datetime.today()

    # sqlite_copy_shelve(filename_shelve, 'test.db', 'temp')
    r = sqlite_search_data('140422 02:00', 'test.db', 'temp')
    # e = json.loads(r)
    print(type(r))
    st = r[2:-3]
    print(st)
    stud_obj = json.loads(st)
    print(stud_obj)
    print(stud_obj['borisov_power_1'])
    # print(stud_obj[''])
    # dictionary = dict(r[2:-2])
    # print(dictionary["soligorsk_water_hot"])
    print(dictionary_registers)

    # help(sqlite_copy_shelve)

    # print(e[0])
    # gas(2022, 4, 13, 11, 6)
    # print_shelve()
    # print(datetime.today() - a)

    help(sqlite_search_data)

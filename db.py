# This is a sample Python script.
from pymodbus.client.sync import ModbusTcpClient
from pymodbus.payload import BinaryPayloadDecoder
from pymodbus.constants import Endian
import shelve
# from pymodbus.payload import BinaryPayloadBuilder
from datetime import datetime
from openpyxl import Workbook, load_workbook

# import schedule

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.
IP_Address = '127.0.0.1'
filename_shelve = 'db'

start = True
run = True
run_gas: bool = True
run_power = True
run_water = True

gas_minute_start = 14
gas_minute_stop = gas_minute_start + 1

# время для общего цикла
start_minute = 0
stop_minute = start_minute + 1

soligorsk_power = 'soligorsk_power'
soligorsk_water_cold = 'soligorsk_water_cold'
soligorsk_water_hot = 'soligorsk_water_hot'
soligorsk_heat = 'soligorsk_heat'

borisov_power_1 = 'borisov_power_1'
borisov_power_2 = 'borisov_power_2'
borisov_water = 'borisov_water'
borisov_heat = 'borisov_heat'

bobruisk_power = 'bobruisk_power'
bobruisk_water = 'bobruisk_water'
bobruisk_heat = 'bobruisk_heat'

minsk_co_gaz_1VR = 'minsk_co_gaz_1VR'
minsk_co_gaz_1VS = 'minsk_co_gaz_1VS'
minsk_co_gaz_1P = 'minsk_co_gaz_1P'
minsk_co_gaz_1T = 'minsk_co_gaz_1T'

dictionary_registers = {soligorsk_power: 0, soligorsk_water_cold: 1,
                        soligorsk_water_hot: 2, soligorsk_heat: 3,

                        borisov_power_1: 4, borisov_power_2: 5, borisov_water: 6, borisov_heat: 7,

                        bobruisk_power: 8, bobruisk_water: 9, bobruisk_heat: 10,

                        minsk_co_gaz_1VR: 11, minsk_co_gaz_1VS: 12, minsk_co_gaz_1P: 13, minsk_co_gaz_1T: 14}
dictionary_temp = {}


def read_registers(dict_reg, dict_temp):
    print('ok')
    client_modbus = ModbusTcpClient(IP_Address)
    dict_temp.clear()
    current_datetime = datetime.today().strftime("%d%m%y %H:%M")
    if client_modbus.connect():
        for key in dict_reg:
            try:
                read_value = client_modbus.read_input_registers(dictionary_registers[key], 2)
                real_decoder = BinaryPayloadDecoder.fromRegisters(read_value.registers, byteorder=Endian.Big,
                                                                  wordorder=Endian.Little)
                value = real_decoder.decode_32bit_float()
                dict_temp[f'{key}'] = value
                # print(f'{key} : {value}')
            except BaseException as e:
                print(f'{e} {key}')

        with shelve.open(filename_shelve) as data:
            data[str(current_datetime)] = dict_temp
            data.close()
        client_modbus.close()
        # print(dict_temp)
    else:
        print(f'not connections {IP_Address}')


def electricity(
        electricity_year=datetime.today().year,
        electricity_month=datetime.today().month,
        electricity_day=datetime.today().day,
        electricity_hour=datetime.today().hour,
        electricity_minute=datetime.today().minute):
    try:
        wb = load_workbook('./electricity/example_el.xlsx')
        ws = wb.active
        with shelve.open(filename_shelve) as states_electricity:
            # запись месяц назад на 28 число 9.00 (28, 9,)
            data_dictionary = states_electricity.get(datetime(electricity_year, electricity_month - 1, electricity_day,
                                                              electricity_hour,
                                                              electricity_minute).strftime("%d%m%y %H:%M"), {0})
            if data_dictionary == {0}:
                print('Для электроэнергии сохранения по этому адресу за прошлый месяц нет - 0')
                ws['D15'].value = 0
                ws['D18'].value = 0
                ws['D19'].value = 0
                ws['D22'].value = 0
            else:
                print('Для электроэнергии заполняем таблицу за прошлый месяц')
                ws['D15'].value = data_dictionary[bobruisk_power]
                ws['D18'].value = data_dictionary[borisov_power_1]
                ws['D19'].value = data_dictionary[borisov_power_2]
                ws['D22'].value = data_dictionary[soligorsk_power]

            data_dictionary.clear()
            # запись за текущий месяц на 28 число 9.00 ( 28, 9,)
            data_dictionary = states_electricity.get(datetime(electricity_year, electricity_month, electricity_day,
                                                              electricity_hour, electricity_minute).strftime(
                "%d%m%y %H:%M"), {0})
            if data_dictionary == {0}:
                print('Для электроэнергии сохранения по этому адресу за текущий месяц нет - 0')
                ws['C15'].value = 0
                ws['C18'].value = 0
                ws['C19'].value = 0
                ws['C22'].value = 0
            else:
                print('Для электроэнергии заполняем таблицу за текущий месяц')
                ws['C15'].value = data_dictionary[bobruisk_power]
                ws['C18'].value = data_dictionary[borisov_power_1]
                ws['C19'].value = data_dictionary[borisov_power_2]
                ws['C22'].value = data_dictionary[soligorsk_power]

            wb.save(f'./electricity/{datetime.today().strftime("%B%Y")}_el.xlsx')  # сохраним в новой книге
            wb.close()
            states_electricity.close()
    except BaseException as e:
        print(e)


def water(
        water_year=datetime.today().year,
        water_month=datetime.today().month,
        water_day=datetime.today().day,
        water_hour=datetime.today().hour,
        water_minute=datetime.today().minute):
    try:
        wb = load_workbook('./water/example_w.xlsx')
        ws = wb.active
        with shelve.open(filename_shelve) as states_water:
            # запись месяц назад на 1 число 9.00
            data_dictionary = states_water.get(datetime(water_year, water_month - 1, water_day, water_hour,
                                                        water_minute).strftime("%d%m%y %H:%M"), {0})
            if data_dictionary == {0}:
                print('Для воды сохранения по этому адресу за прошлый месяц нет - 0')
                ws['D9'].value = 0
                ws['D10'].value = 0
                ws['D11'].value = 0
                ws['D12'].value = 0
            else:
                print('Для воды заполняем таблицу за прошлый месяц')
                ws['D9'].value = data_dictionary[bobruisk_water]
                ws['D10'].value = data_dictionary[borisov_water]
                ws['D11'].value = data_dictionary[soligorsk_water_cold]
                ws['D12'].value = data_dictionary[soligorsk_water_hot]

            data_dictionary.clear()
            data_dictionary = states_water.get(datetime(water_year, water_month, water_day, water_hour,
                                                        water_minute).strftime("%d%m%y %H:%M"), {0})
            if data_dictionary == {0}:
                print('Для воды сохранения по этому адресу за текущий месяц нет - 0')
                ws['E9'].value = 0
                ws['E10'].value = 0
                ws['E11'].value = 0
                ws['E12'].value = 0
            else:
                print('Для воды заполняем таблицу за текущий месяц')
                ws['E9'].value = data_dictionary[bobruisk_water]
                ws['E10'].value = data_dictionary[borisov_water]
                ws['E11'].value = data_dictionary[soligorsk_water_cold]
                ws['E12'].value = data_dictionary[soligorsk_water_hot]

            wb.save(f'./water/{datetime.today().strftime("%B%Y")}_w.xlsx')  # сохраним в новой книге
            wb.close()
            states_water.close()
    except BaseException as e:
        print(e)


def heat(
        heat_year=datetime.today().year,
        heat_month=datetime.today().month,
        heat_day=datetime.today().day,
        heat_hour=datetime.today().hour,
        heat_minute=datetime.today().minute):
    try:
        wb = load_workbook('./heat/example_h.xlsx')
        ws = wb.active
        with shelve.open(filename_shelve) as states_heat:
            # запись месяц назад на 1 число 9.00
            data_dictionary = states_heat.get(datetime(heat_year, heat_month - 1, heat_day, heat_hour,
                                                       heat_minute).strftime("%d%m%y %H:%M"), {0})
            if data_dictionary == {0}:
                print('Для тепла сохранения по этому адресу за прошлый месяц нет - 0')
                ws['D9'].value = 0
                ws['D10'].value = 0
                ws['D11'].value = 0
                ws['D12'].value = 0
            else:
                print('Для тепла заполняем таблицу за прошлый месяц')
                ws['D9'].value = data_dictionary[bobruisk_heat]
                ws['D10'].value = data_dictionary[borisov_water]
                ws['D11'].value = data_dictionary[soligorsk_water_cold]
                ws['D12'].value = data_dictionary[soligorsk_water_hot]

            data_dictionary.clear()
            data_dictionary = states_heat.get(datetime(heat_year, heat_month, heat_day, heat_hour,
                                                       heat_minute).strftime("%d%m%y %H:%M"), {0})
            if data_dictionary == {0}:
                print('Для тепла сохранения по этому адресу за текущий месяц нет - 0')
                ws['E9'].value = 0
                ws['E10'].value = 0
                ws['E11'].value = 0
                ws['E12'].value = 0

            else:
                print('Для тепла заполняем таблицу за текущий месяц')
                ws['E9'].value = data_dictionary[bobruisk_heat]
                ws['E10'].value = data_dictionary[borisov_water]
                ws['E11'].value = data_dictionary[soligorsk_water_cold]
                ws['E12'].value = data_dictionary[soligorsk_water_hot]

            wb.save(f'./heat/{datetime.today().strftime("%B%Y")}_h.xlsx')  # сохраним в новой книге
            wb.close()
            states_heat.close()
    except BaseException as e:
        print(e)


def gas(
        gas_year=datetime.today().year,
        gas_month=datetime.today().month,
        gas_day=datetime.today().day,
        gas_hour=datetime.today().hour,
        gas_minute=datetime.today().minute):
    try:
        wb = load_workbook('./gas/example_gas.xlsx')
        ws = wb.active
        with shelve.open(filename_shelve) as states:
            # запись месяц назад на 1 число 9.00 ( 1, 9,)
            data_dictionary = states.get(datetime(gas_year, gas_month, gas_day, gas_hour,
                                                  gas_minute).strftime("%d%m%y %H:%M"), {0})
            if data_dictionary == {0}:
                print('Для газа сохранения по этому адресу нет - 0')
                ws['C5'].value = 0
                ws['D5'].value = 0
                ws['E5'].value = 0
                ws['F5'].value = 0
                ws['H5'].value = 0
                ws['I5'].value = 0
            else:
                print('Для газа заполняем таблицу')
                ws['C5'].value = data_dictionary[minsk_co_gaz_1VS]
                ws['D5'].value = data_dictionary[minsk_co_gaz_1VR]
                ws['E5'].value = data_dictionary[minsk_co_gaz_1P]
                ws['F5'].value = data_dictionary[minsk_co_gaz_1T]
                ws['H5'].value = data_dictionary[minsk_co_gaz_1VS]
                ws['I5'].value = data_dictionary[minsk_co_gaz_1VR]

            data_dictionary.clear()
            wb.save(f'./gas/{datetime.today().strftime("%B%Y")}_gas.xlsx')  # сохраним в новой книге
            wb.close()
            states.close()
    except BaseException as e:
        print(e)


def run_report_gas(
        function_report,
        run_report_function,
        function_year=datetime.today().year,
        function_month=datetime.today().month,
        function_day=datetime.today().day,
        function_hour=datetime.today().hour,
        function_minute=datetime.today().minute,

        report_day=datetime.today().day,
        report_hour=datetime.today().hour,
        report_minute=datetime.today().minute,
        report_minute_stop=0):
    if not hasattr(run_report_gas, '_state'):  # инициализация значения
        run_report_gas._state = True
    # global run_gas
    # print(f'{report_day}, {report_hour}, {report_minute}')
    now_report = datetime.today()
    # print(f'вход в сравнение, run {run_report_function}')

    if now_report.day == report_day and \
            now_report.hour == report_hour and \
            now_report.minute == report_minute and \
            run_report_gas._state:

        run_report_gas._state = False
        function_report(function_year, function_month, function_day, function_hour, function_minute)
        # print(f'функция работает, run {run_gas}')

    elif now_report.minute > report_minute_stop:
        run_report_gas._state = True
        # print(f'функция не работает, run {run_gas}')


def run_report_electricity(
        function_report,
        run_report_function,
        function_year=datetime.today().year,
        function_month=datetime.today().month,
        function_day=datetime.today().day,
        function_hour=datetime.today().hour,
        function_minute=datetime.today().minute,

        report_day=datetime.today().day,
        report_hour=datetime.today().hour,
        report_minute=datetime.today().minute,
        report_minute_stop=0):
    if not hasattr(run_report_electricity, '_state'):  # инициализация значения
        run_report_electricity._state = True
    # global run_gas
    # print(f'{report_day}, {report_hour}, {report_minute}')
    now_report = datetime.today()
    # print(f'вход в сравнение, run {run_report_function}')

    if now_report.day == report_day and \
            now_report.hour == report_hour and \
            now_report.minute == report_minute and \
            run_report_electricity._state:

        run_report_electricity._state = False
        function_report(function_year, function_month, function_day, function_hour, function_minute)
        # print(f'функция работает, run {run_gas}')

    elif now_report.minute > report_minute_stop:
        run_report_electricity._state = True
        # print(f'функция не работает, run {run_gas}')


def run_report_water(
        function_report,
        run_report_function,
        function_year=datetime.today().year,
        function_month=datetime.today().month,
        function_day=datetime.today().day,
        function_hour=datetime.today().hour,
        function_minute=datetime.today().minute,

        report_day=datetime.today().day,
        report_hour=datetime.today().hour,
        report_minute=datetime.today().minute,
        report_minute_stop=0):
    if not hasattr(run_report_water, '_state'):  # инициализация значения
        run_report_water._state = True
    # global run_gas
    # print(f'{report_day}, {report_hour}, {report_minute}')
    now_report = datetime.today()
    # print(f'вход в сравнение, run {run_report_function}')

    if now_report.day == report_day and \
            now_report.hour == report_hour and \
            now_report.minute == report_minute and \
            run_report_water._state:

        run_report_water._state = False
        function_report(function_year, function_month, function_day, function_hour, function_minute)
        # print(f'функция работает, run {run_gas}')

    elif now_report.minute > report_minute_stop:
        run_report_water._state = True
        # print(f'функция не работает, run {run_gas}')


def run_report_heat(
        function_report,
        run_report_function,
        function_year=datetime.today().year,
        function_month=datetime.today().month,
        function_day=datetime.today().day,
        function_hour=datetime.today().hour,
        function_minute=datetime.today().minute,

        report_day=datetime.today().day,
        report_hour=datetime.today().hour,
        report_minute=datetime.today().minute,
        report_minute_stop=0):
    if not hasattr(run_report_heat, '_state'):  # инициализация значения
        run_report_heat._state = True
    # global run_gas
    # print(f'{report_day}, {report_hour}, {report_minute}')
    now_report = datetime.today()
    # print(f'вход в сравнение, run {run_report_function}')

    if now_report.day == report_day and \
            now_report.hour == report_hour and \
            now_report.minute == report_minute and \
            run_report_heat._state:

        run_report_heat._state = False
        function_report(function_year, function_month, function_day, function_hour, function_minute)
        # print(f'функция работает, run {run_gas}')

    elif now_report.minute > report_minute_stop:
        run_report_heat._state = True
        # print(f'функция не работает, run {run_gas}')


if __name__ == '__main__':

    print(f'Старт опроса - {datetime.today()}')

    while start:

        run_report_gas(gas, run_gas, 2022, 4, 13, 15, 0,
                       14, 23, gas_minute_start, gas_minute_stop)

        run_report_heat(heat, run_gas, 2022, 4, 13, 15, 0,
                        14, 23, gas_minute_start, gas_minute_stop)

        run_report_electricity(electricity, run_gas, 2022, 4, 13, 15, 0,
                               14, 23, gas_minute_start, gas_minute_stop)

        run_report_water(water, run_gas, 2022, 4, 13, 15, 0,
                         14, 23, gas_minute_start, gas_minute_stop)

        now = datetime.today()
        deadline = datetime(datetime.today().year, datetime.today().month, datetime.today().day, datetime.today().hour,
                            start_minute)

        if now.day == deadline.day and now.hour == deadline.hour and now.minute == deadline.minute and run:
            run = False
            read_registers(dictionary_registers, dictionary_temp)

            with shelve.open(filename_shelve) as states:
                for key in states:
                    print(key, " - ", states[key])
                    print('-----------------------')
            states.close()
            # print(run)

        elif now.minute > stop_minute:
            run = True
        print('Hello')
